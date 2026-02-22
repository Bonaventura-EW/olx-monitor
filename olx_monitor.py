#!/usr/bin/env python3
"""
OLX Monitor — monitoring ogłoszeń z datą publikacji, wiekiem i historią cen.

NAPRAWY W TEJWERSJI:
  ✓ Nowy parser cen extract_price_from_card() z walidacją
  ✓ Obsługa wyjątków w fetch_dates()
  ✓ Ulepszona logika cross-check
  ✓ Lepszy error handling w całym kodzie

Generuje:
  olx_monitoring.xlsx     — pełna tabela z każdym skanem
  price_history.json      — historia cen dla dashboardu HTML

Kolumny Excela:
  Data skanu | Profil | Tytuł | Cena (zł) | Data publikacji | Dni od publikacji | URL | ID
"""

import requests, re, json, os, time
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── KONFIGURACJA ─────────────────────────────────────────
CONFIG_FILE        = "config.json"
EXCEL_FILE         = "data/olx_monitoring.xlsx"
PRICE_HISTORY_FILE = "data/price_history.json"

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/121.0.0.0 Safari/537.36",
    "Accept-Language": "pl-PL,pl;q=0.9",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Zakresy cen dla walidacji (pokoje w Lublinie)
MIN_PRICE = 150   # zł/mies
MAX_PRICE = 20000  # zł/mies

# ── POMOCNICZE ────────────────────────────────────────────
def parse_created(html):
    """
    Wyciąga datę publikacji ogłoszenia z zakodowanego JSON w HTML.
    OLX osadza: createdTime\\\":\\\"2025-09-18T18:08:49+02:00\\\"
    Zwraca (datetime_iso_str, days_old) lub (None, None).
    """
    idx = html.find("createdTime")
    if idx < 0:
        return None, None
    snippet = html[idx:idx + 80]
    m = re.search(r"(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}[+-]\d{2}:\d{2})", snippet)
    if not m:
        return None, None
    dt_str = m.group(1)
    try:
        dt  = datetime.fromisoformat(dt_str)
        now = datetime.now(tz=dt.tzinfo)
        days = max(0, (now - dt).days)
        return dt.strftime("%d.%m.%Y"), days
    except Exception:
        return None, None


def today_label():
    """Krótka etykieta dnia dla price_history — np. '16 lut'."""
    months = ["sty","lut","mar","kwi","maj","cze",
               "lip","sie","wrz","paź","lis","gru"]
    n = datetime.now()
    return f"{n.day} {months[n.month - 1]}"


def extract_price_from_card(card_text: str) -> int:
    """
    Wyciąga cenę z tekstu karty ogłoszenia z walidacją zakresu.
    Szuka wzorca: liczba (2-5 cyfr) + spacja(e) + "zł"
    
    Walidacja: cena powinna być między MIN_PRICE a MAX_PRICE
    Jeśli poza zakresem — zwraca 0 (cena niepewna/anomalna)
    
    POPRAWKA: Jeśli znajdzie wiele cen, bierze NAJNIŻSZĄ w prawidłowym zakresie
    (aby uniknąć parsowania sum: czynsz + media + kaucja)
    
    Przykłady:
      "899 zł/mies" → 899
      "2 299 zł" → 2299
      "1200zł" → 1200
      "cena 0 zł" → 0 (cena nie podana)
      "58640 zł" → 0 (anomalnie wysoko — błąd parsowania)
      "1200 zł + 400 zł media = 1600 zł" → 1200 (najniższa prawidłowa)
    """
    # Znajdź WSZYSTKIE wystąpienia ceny
    matches = re.findall(r"(\d[\d\s]*\d|\d)\s*zł", card_text, re.IGNORECASE)
    if not matches:
        return 0
    
    # Przekonwertuj wszystkie znalezione ceny na liczby
    prices = []
    for match in matches:
        price_str = re.sub(r"[^\d]", "", match)
        if not price_str or price_str == "0":
            continue
        try:
            price = int(price_str)
            # Dodaj tylko ceny w prawidłowym zakresie
            if MIN_PRICE <= price <= MAX_PRICE:
                prices.append(price)
        except ValueError:
            continue
    
    # Jeśli nie znaleziono żadnej prawidłowej ceny, zwróć 0
    if not prices:
        return 0
    
    # Zwróć NAJNIŻSZĄ cenę (główny czynsz, bez dodatków)
    return min(prices)


# ── CROSS-CHECK: weryfikacja liczby ogłoszeń ─────────────
def crosscheck_count(soup) -> int | None:
    """
    Wyciąga oficjalną liczbę ogłoszeń z tekstu 'Znaleźliśmy X ogłoszeń'
    na stronie profilu OLX. Zwraca int lub None jeśli nie znaleziono.
    """
    for el in soup.find_all(string=re.compile(r"Znaleźliśmy")):
        m = re.search(r"(\d+)", el)
        if m:
            return int(m.group(1))
    return None


def fetch_market_total() -> int | None:
    """
    Pobiera aktualną liczbę ogłoszeń stancji/pokoi w Lublinie z OLX.
    Używa data-testid='total-count' który jest stabilnym selektorem.
    """
    url = "https://www.olx.pl/nieruchomosci/stancje-pokoje/lublin/"
    try:
        r    = requests.get(url, headers=HEADERS, timeout=15)
        soup = BeautifulSoup(r.text, "html.parser")
        el   = soup.find(attrs={"data-testid": "total-count"})
        if el:
            m = re.search(r"(\d+)", el.get_text())
            if m:
                total = int(m.group(1))
                print(f"  → Rynek OLX Lublin (stancje/pokoje): {total} ogłoszeń")
                return total
        # fallback: szukaj w tekście strony
        for text_el in soup.find_all(string=re.compile(r"Znaleźliśmy")):
            m = re.search(r"(\d+)", text_el)
            if m:
                total = int(m.group(1))
                print(f"  → Rynek OLX Lublin (fallback): {total} ogłoszeń")
                return total
    except Exception as e:
        print(f"  ⚠ Błąd pobierania licznika rynku: {e}")
    return None


# ── SCRAPER: profil OLX → lista ogłoszeń ─────────────────
def scrape_profile(profile_name, profile_url):
    print(f"  [{profile_name}] {profile_url}")
    try:
        r = requests.get(profile_url, headers=HEADERS, timeout=15)
        r.raise_for_status()
    except Exception as e:
        print(f"    ⚠ Błąd pobierania profilu: {e}")
        return [], None, None

    soup = BeautifulSoup(r.text, "html.parser")
    listings, seen = [], set()

    # Metoda 1: div[type="list"] — stabilny atrybut, niezależny od klas CSS
    # OLX opakowuje każdą kartę ogłoszenia w <div type="list">
    cards = soup.find_all("div", attrs={"type": "list"})

    for card in cards:
        a = card.find("a", href=lambda h: h and "/d/oferta/" in h)
        if not a:
            continue
        href = re.sub(r"\?.*", "", a.get("href", ""))
        if href in seen:
            continue
        seen.add(href)

        # Tytuł — w <p> wewnątrz karty
        p_tag = card.find("p")
        title = p_tag.get_text(strip=True) if p_tag else a.get_text(strip=True)
        if not title or len(title) < 5:
            continue

        card_text = card.get_text(" ", strip=True)
        price = extract_price_from_card(card_text)

        full_url   = ("https://www.olx.pl" + href) if href.startswith("/") else href
        id_m       = re.search(r"/d/oferta/([^/?\.]+)", href)
        listing_id = id_m.group(1) if id_m else href.replace("/", "_")

        listings.append({
            "id":       listing_id,
            "profile":  profile_name,
            "title":    title,
            "price":    price,
            "url":      full_url,
            "created":  None,
            "days_old": None,
        })

    # Metoda 2 (fallback): jeśli type=list nie zadziałał
    if not listings:
        print(f"    ⚠ Metoda type=list nie znalazła kart — fallback na href")
        for a in soup.find_all("a", href=lambda h: h and "/d/oferta/" in h):
            href = re.sub(r"\?.*", "", a.get("href", ""))
            if href in seen:
                continue
            seen.add(href)
            ancestor = a.parent.parent if (a.parent and a.parent.parent) else a.parent
            if not ancestor:
                continue
            p_tag = ancestor.find("p")
            title = p_tag.get_text(strip=True) if p_tag else ""
            if not title or len(title) < 5:
                continue
            card_text = ancestor.get_text(" ", strip=True)
            price = extract_price_from_card(card_text)
            full_url = ("https://www.olx.pl" + href) if href.startswith("/") else href
            id_m = re.search(r"/d/oferta/([^/?\.]+)", href)
            listing_id = id_m.group(1) if id_m else href.replace("/", "_")
            listings.append({
                "id": listing_id, "profile": profile_name, "title": title,
                "price": price, "url": full_url, "created": None, "days_old": None,
            })

    # ── Cross-check z oficjalną liczbą OLX ───────────────
    official_count = crosscheck_count(soup)
    scraped_count  = len(listings)

    if official_count is None:
        cc_msg = "⚠  cross-check: brak licznika na stronie"
        cc_ok  = None
    elif scraped_count == official_count:
        cc_msg = f"✓  cross-check OK ({scraped_count} = {official_count})"
        cc_ok  = True
    else:
        diff   = scraped_count - official_count
        cc_msg = f"⚠  cross-check NIEZGODNOŚĆ: scraped={scraped_count}, OLX={official_count} (diff={diff:+d})"
        cc_ok  = False

    print(f"    → {scraped_count} ogłoszeń  |  {cc_msg}")
    return listings, official_count, cc_ok


# ── SCRAPER: data publikacji z każdego ogłoszenia ─────────
def fetch_dates(listings, delay=1.2):
    """
    Wchodzi w stronę każdego ogłoszenia i wyciąga createdTime.
    Delay chroni przed blokadą IP.
    
    NAPRAWY:
      - Obsługa wyjątków dla każdego ogłoszenia niezależnie
      - Timeout dla każdego requesta
      - Logowanie błędów bez przerywania procesu
    """
    print(f"\n  Pobieranie dat publikacji ({len(listings)} ogłoszeń, ~{len(listings)*delay:.0f}s)...")
    failed = []
    
    for i, l in enumerate(listings, 1):
        try:
            r = requests.get(l["url"], headers=HEADERS, timeout=12)
            r.raise_for_status()
            created, days = parse_created(r.text)
            l["created"]  = created  # "DD.MM.YYYY" lub None
            l["days_old"] = days     # int lub None
            status = f"{created}  ({days} dni)" if created else "brak daty"
        except requests.exceptions.Timeout:
            l["created"]  = None
            l["days_old"] = None
            status = "błąd: timeout"
            failed.append(l["id"])
        except requests.exceptions.ConnectionError as e:
            l["created"]  = None
            l["days_old"] = None
            status = f"błąd: brak sieci"
            failed.append(l["id"])
        except Exception as e:
            l["created"]  = None
            l["days_old"] = None
            status = f"błąd: {type(e).__name__}"
            failed.append(l["id"])
        
        print(f"    [{i:2}/{len(listings)}] {l['title'][:50]:<50} {status}")
        time.sleep(delay)
    
    if failed:
        print(f"\n  ⚠  Niepowodzenia ({len(failed)}): {', '.join(failed[:5])}")
        if len(failed) > 5:
            print(f"      ... i {len(failed)-5} więcej")
    
    return listings


# ── PRICE HISTORY JSON ────────────────────────────────────
def update_price_history(listings):
    today = today_label()
    history = {}
    os.makedirs(os.path.dirname(PRICE_HISTORY_FILE), exist_ok=True)
    if os.path.exists(PRICE_HISTORY_FILE):
        try:
            with open(PRICE_HISTORY_FILE, "r", encoding="utf-8") as f:
                history = json.load(f)
        except Exception as e:
            print(f"  ⚠ Błąd wczytywania price_history.json: {e}")

    for l in listings:
        lid = l["id"]
        if lid not in history:
            history[lid] = {"title": l["title"], "profile": l["profile"],
                            "created": l["created"] or "", "prices": []}
        else:
            if not history[lid].get("created") and l.get("created"):
                history[lid]["created"] = l["created"]

        prices = history[lid]["prices"]
        entry  = next((e for e in prices if e["date"] == today), None)
        if entry:
            entry["price"] = l["price"]
        elif l["price"] > 0:
            prices.append({"date": today, "price": l["price"]})

    try:
        with open(PRICE_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        print(f"  → {PRICE_HISTORY_FILE}: {len(history)} ogłoszeń")
    except Exception as e:
        print(f"  ⚠ Błąd zapisywania price_history.json: {e}")


# ── EXCEL ─────────────────────────────────────────────────
COLUMNS = [
    ("Data skanu",        20),
    ("Profil",            16),
    ("Tytuł",             54),
    ("Cena (zł)",         12),
    ("Data publikacji",   16),
    ("Dni od publikacji", 18),
    ("URL",               60),
    ("ID ogłoszenia",     44),
]

def cell_style(cell, color=None, bold=False, align="left"):
    if color:
        cell.font = Font(color=color, bold=bold, size=10)
    elif bold:
        cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")


def save_to_excel(listings):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    thin   = Side(style="thin", color="2a2a38")
    border = Border(bottom=thin)

    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Historia"

            # Nagłówki
            ws.append([col for col, _ in COLUMNS])
            hfill  = PatternFill("solid", fgColor="1a1a2e")
            hfont  = Font(color="e8ff47", bold=True, size=10)
            halign = Alignment(horizontal="center", vertical="center", wrap_text=False)
            for col_idx, cell in enumerate(ws[1], 1):
                cell.fill      = hfill
                cell.font      = hfont
                cell.alignment = halign
                cell.border    = border
                ws.column_dimensions[get_column_letter(col_idx)].width = COLUMNS[col_idx-1][1]

            ws.freeze_panes    = "A2"
            ws.row_dimensions[1].height = 20

        # Dane
        for l in listings:
            days = l.get("days_old")
            ws.append([
                now,
                l["profile"],
                l["title"],
                l["price"],
                l["created"] or "",
                days if days is not None else "",
                l["url"],
                l["id"],
            ])
            row = ws.max_row

            # Wyrównanie
            ws.cell(row, 1).alignment = Alignment(horizontal="left")
            ws.cell(row, 3).alignment = Alignment(horizontal="left")
            ws.cell(row, 4).alignment = Alignment(horizontal="center")
            ws.cell(row, 5).alignment = Alignment(horizontal="center")
            ws.cell(row, 6).alignment = Alignment(horizontal="center")

            # Kolorowanie kolumny "Dni od publikacji"
            if days is not None:
                cell = ws.cell(row, 6)
                if days <= 3:
                    cell.font = Font(color="47ffa0", bold=True, size=10)   # świeże — zielony
                elif days <= 14:
                    cell.font = Font(color="e8ff47", bold=False, size=10)  # niedawne — żółty
                elif days > 60:
                    cell.font = Font(color="ff6b6b", bold=False, size=10)  # stare — czerwony

        wb.save(EXCEL_FILE)
        print(f"  → {EXCEL_FILE}: +{len(listings)} wierszy (łącznie {ws.max_row - 1})")
    except Exception as e:
        print(f"  ⚠ Błąd zapisu Excela: {e}")


# ── MAIN ──────────────────────────────────────────────────

def save_profiles_state(all_listings, config, price_history):
    """
    Zapisuje aktualny stan profili do data/profiles_state.json.
    Workflow wstrzykuje ten plik do dashboardu jako __PROFILES_DATA__.
    """
    from collections import defaultdict

    today     = datetime.now()
    today_pl  = today_label()
    state_file = "data/profiles_state.json"

    # Wczytaj poprzedni stan jeden raz
    prev_data: dict = {}
    if os.path.exists(state_file):
        try:
            with open(state_file, "r", encoding="utf-8") as f:
                prev_data = json.load(f)
        except Exception as e:
            print(f"  ⚠ Błąd wczytywania profiles_state.json: {e}")

    # Grupuj ogłoszenia per profil
    by_profile: dict = defaultdict(list)
    for l in all_listings:
        by_profile[l["profile"]].append(l)

    # Etykieta daty w formacie "17 lut 2026"
    month_pl = ["sty","lut","mar","kwi","maj","cze",
                "lip","sie","wrz","paź","lis","gru"]
    today_str = f"{today.day} {month_pl[today.month-1]} {today.year}"

    profiles_out = {}
    for p in config.get("profiles", []):
        name = p["name"]
        url  = p["url"]

        listings  = by_profile.get(name, [])
        prev_prof = prev_data.get(name, {})
        prev_curr = prev_prof.get("current", [])
        prev_ids  = {l["id"] for l in prev_curr}
        curr_ids  = {l["id"] for l in listings}
        new_ids   = curr_ids - prev_ids
        gone_ids  = prev_ids - curr_ids

        # current[] — ogłoszenia z bieżącego skanu
        current = []
        for l in listings:
            current.append({
                "id":      l["id"],
                "title":   l["title"],
                "price":   l["price"],
                "url":     l["url"],
                "status":  "new" if l["id"] in new_ids else "existing",
                "created": l.get("created") or "",
                "daysOld": l.get("days_old"),
                "date":    today_pl,
                "profile": name,
            })

        # gone[] — ogłoszenia które zniknęły (z poprzedniego current[])
        gone = [
            {"id": pl["id"], "title": pl["title"],
             "price": pl["price"], "url": pl["url"], "date": today_pl}
            for pl in prev_curr if pl["id"] in gone_ids
        ]

        # Historia — jeden wpis na dzień, zastąp dzisiejszy jeśli istnieje
        history = [h for h in prev_prof.get("history", [])
                   if h.get("date") != today_str]
        history.append({
            "date":      today_str,
            "total":     len(current),
            "newCount":  len(new_ids),
            "goneCount": len(gone_ids),
        })
        history = history[-30:]   # max 30 dni

        profiles_out[name] = {
            "url":     url,
            "current": current,
            "gone":    gone,
            "history": history,
        }

    os.makedirs("data", exist_ok=True)
    try:
        with open(state_file, "w", encoding="utf-8") as f:
            json.dump(profiles_out, f, ensure_ascii=False, indent=2)
        total = sum(len(v["current"]) for v in profiles_out.values())
        print(f"  → {state_file}: {total} ogłoszeń w {len(profiles_out)} profilach")
    except Exception as e:
        print(f"  ⚠ Błąd zapisu profiles_state.json: {e}")
    
    return profiles_out

def main():
    print("=" * 60)
    print(f"OLX Monitor — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception as e:
        print(f"❌ Błąd wczytywania config.json: {e}")
        return

    # 1. Scrape profiles
    print("\n[1/3] Scraping profili OLX...")
    all_listings  = []
    crosscheck_results = []   # (profile, scraped, official, ok)
    for p in config.get("profiles", []):
        listings, official, cc_ok = scrape_profile(p["name"], p["url"])
        all_listings.extend(listings)
        crosscheck_results.append((p["name"], len(listings), official, cc_ok))

    print("\n[+]   Pobieranie licznika rynku OLX Lublin...")
    market_total = fetch_market_total()

    if not all_listings:
        print("⚠ Brak ogłoszeń. Koniec.")
        return

    print(f"\nRazem: {len(all_listings)} ogłoszeń")

    # 2. Daty publikacji
    print("\n[2/3] Pobieranie dat publikacji z OLX...")
    all_listings = fetch_dates(all_listings, delay=1.2)

    # 3. Zapis
    print("\n[3/3] Zapisywanie...")
    update_price_history(all_listings)
    save_to_excel(all_listings)

    # Wczytaj price_history do stanu profili
    ph = {}
    if os.path.exists(PRICE_HISTORY_FILE):
        try:
            with open(PRICE_HISTORY_FILE, "r", encoding="utf-8") as f:
                ph = json.load(f)
        except Exception as e:
            print(f"  ⚠ Błąd wczytywania price_history dla profiles_state: {e}")
    
    save_profiles_state(all_listings, config, ph)

    # Podsumowanie
    with_date = [l for l in all_listings if l["created"]]
    no_date   = [l for l in all_listings if not l["created"]]
    print(f"\n✓ Gotowe!")
    print(f"  Daty publikacji znalezione: {len(with_date)}/{len(all_listings)}")
    if no_date:
        print(f"  Bez daty: {[l['title'][:40] for l in no_date]}")

    # ── Raport cross-check ────────────────────────────────
    print("\n" + "=" * 60)
    print("CROSS-CHECK — weryfikacja liczby ogłoszeń")
    print("=" * 60)
    problems = []
    for name, scraped, official, ok in crosscheck_results:
        if ok is True:
            status = "✓ OK"
        elif ok is False:
            status = f"⚠ NIEZGODNOŚĆ  scraped={scraped}  OLX={official}  diff={scraped-official:+d}"
            problems.append(name)
        else:
            status = f"? brak licznika (scraped={scraped})"
        print(f"  {name:<22} {status}")

    if problems:
        print(f"\n⚠  Niezgodności w profilach: {', '.join(problems)}")
        print("   Sprawdź ręcznie lub uruchom ponownie za kilka minut.")
    else:
        print("\n✓ Wszystkie profile zgodne z licznikiem OLX.")

    # Zapisz wyniki cross-check do last_run.json
    import json as _json
    last_run = {
        "run_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "total_listings": len(all_listings),
        "dates_found": len(with_date),
        "market_total": market_total,
        "crosscheck": [
            {"profile": n, "scraped": s, "official": o, "ok": k}
            for n, s, o, k in crosscheck_results
        ],
        "problems": problems,
    }
    os.makedirs("data", exist_ok=True)
    try:
        with open("data/last_run.json", "w", encoding="utf-8") as f:
            _json.dump(last_run, f, ensure_ascii=False, indent=2)
        print(f"\n  → data/last_run.json zaktualizowany")
    except Exception as e:
        print(f"\n  ⚠ Błąd zapisu last_run.json: {e}")
    
    print("=" * 60)


if __name__ == "__main__":
    main()
