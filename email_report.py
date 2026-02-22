"""
Tygodniowy raport e-mail z analizą AI.
Wysyłany w każdy poniedziałek – zbiera dane z ostatnich 7 dni z pliku Excel
i wysyła podsumowanie przez Gmail SMTP.
"""

import smtplib
import json
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
import openpyxl
import requests

# ─── KONFIGURACJA ────────────────────────────────────────────────────────────

SENDER_EMAIL    = "slowholidays00@gmail.com"
RECIPIENT_EMAIL = "malczarski@gmail.com"
EXCEL_FILE      = "data/olx_monitoring.xlsx"

# Nazwy zakładek w Excelu — muszą być identyczne z config.json
PROFILES = [
    "artymiuk",
    "poqui",
    "pokojewlublinie",
    "villahome",
    "dawnypatron",
]

# ─── ZBIERANIE DANYCH Z EXCELA ───────────────────────────────────────────────

def get_weekly_data() -> dict:
    """
    Odczytuje dane z ostatnich 7 dni z Excela - zakładka "Historia".
    - Grupuje po profilu i dacie (jeden rekord na dzień/profil)
    - Posortowane od najnowszej do najstarszej daty
    """
    if not os.path.exists(EXCEL_FILE):
        print(f"⚠  Brak pliku Excel: {EXCEL_FILE}")
        return {}

    wb       = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    week_ago = datetime.now() - timedelta(days=7)
    data     = {}

    # Excel ma jedną zakładkę "Historia" z wszystkimi profilami
    ws_name = "Historia"
    if ws_name not in wb.sheetnames:
        print(f"  ⚠  Brak zakładki '{ws_name}' w Excelu – pomijam")
        return {}

    ws = wb[ws_name]
    print(f"  ✓  Czytam zakładkę: {ws_name}")

    # Słownik: profil -> { data_str -> ostatni rekord z tego dnia }
    profile_data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:  # Data skanu, Profil
            continue
        try:
            # row[0] = Data skanu, row[1] = Profil
            row_dt  = datetime.strptime(str(row[0])[:16], "%Y-%m-%d %H:%M")
            profile = str(row[1])
        except Exception:
            continue

        if row_dt < week_ago:
            continue

        if profile not in profile_data:
            profile_data[profile] = {}

        date_str = row_dt.strftime("%Y-%m-%d")

        # Dla każdego profilu/dnia liczymy statystyki z price_history i profiles_state
        # Tutaj tylko zbieramy daty które wystąpiły
        if date_str not in profile_data[profile]:
            profile_data[profile][date_str] = {
                "_dt":     row_dt,
                "date":    date_str,
                "status":  "OK",
            }
        else:
            # Zachowaj najnowszy timestamp z danego dnia
            if row_dt > profile_data[profile][date_str]["_dt"]:
                profile_data[profile][date_str]["_dt"] = row_dt

    # Oblicz statystyki z profiles_state.json
    if os.path.exists("data/profiles_state.json"):
        with open("data/profiles_state.json", "r", encoding="utf-8") as f:
            profiles_state = json.load(f)

        for profile, dates_dict in profile_data.items():
            if profile not in profiles_state:
                continue

            prof_data = profiles_state[profile]
            history   = prof_data.get("history", [])

            # Mapowanie: data_str -> history entry
            history_map = {}
            for h in history:
                # Format "17 lut 2026" -> "2026-02-17"
                try:
                    date_parts = h["date"].split()
                    if len(date_parts) == 3:
                        day = int(date_parts[0])
                        month_map = {
                            'sty': 1, 'lut': 2, 'mar': 3, 'kwi': 4,
                            'maj': 5, 'cze': 6, 'lip': 7, 'sie': 8,
                            'wrz': 9, 'paź': 10, 'lis': 11, 'gru': 12
                        }
                        month = month_map.get(date_parts[1].lower(), 1)
                        year  = int(date_parts[2])
                        norm_date = f"{year:04d}-{month:02d}-{day:02d}"
                        history_map[norm_date] = h
                except Exception:
                    pass

            # Uzupełnij statystyki
            for date_str, entry in dates_dict.items():
                if date_str in history_map:
                    h = history_map[date_str]
                    entry["total"]   = h.get("total", 0)
                    entry["new"]     = h.get("newCount", 0)
                    entry["deleted"] = h.get("goneCount", 0)
                    entry["net"]     = entry["new"] - entry["deleted"]
                else:
                    entry["total"]   = 0
                    entry["new"]     = 0
                    entry["deleted"] = 0
                    entry["net"]     = 0

    # Przekształć do formatu wyjściowego
    for profile, dates_dict in profile_data.items():
        rows = sorted(dates_dict.values(), key=lambda x: x["_dt"], reverse=True)
        # Usuń pomocnicze pole _dt
        for r in rows:
            del r["_dt"]
        data[profile] = rows
        print(f"  ✓  {profile}: {len(rows)} dni, najnowszy: {rows[0]['date']}")

    return data


def compute_summary(weekly_data: dict) -> dict:
    summary = {}
    for profile, rows in weekly_data.items():
        total_new     = sum(r.get("new", 0)     for r in rows)
        total_deleted = sum(r.get("deleted", 0) for r in rows)
        last_total    = rows[0].get("total", 0)  if rows else 0
        first_total   = rows[-1].get("total", 0) if rows else 0
        errors        = sum(1 for r in rows if r.get("status", "OK") != "OK")

        summary[profile] = {
            "days_tracked":  len(rows),
            "total_new":     total_new,
            "total_deleted": total_deleted,
            "net_week":      total_new - total_deleted,
            "last_count":    last_total,
            "first_count":   first_total,
            "errors":        errors,
            "rows":          rows,
        }
    return summary


# ─── HTML E-MAIL ─────────────────────────────────────────────────────────────

def build_html_email(summary: dict, weekly_data: dict, analysis: str) -> str:
    today      = datetime.now().strftime("%d.%m.%Y")
    week_start = (datetime.now() - timedelta(days=6)).strftime("%d.%m.%Y")

    summary_rows = ""
    for profile, s in summary.items():
        trend     = "↑" if s["net_week"] > 0 else ("↓" if s["net_week"] < 0 else "→")
        new_style = "color:#1a7a3c;font-weight:bold;" if s["total_new"] > 0 else ""
        del_style = "color:#c0392b;font-weight:bold;" if s["total_deleted"] > 0 else ""
        net_color = "#1a7a3c" if s["net_week"] > 0 else ("#c0392b" if s["net_week"] < 0 else "#555")
        err_style = "color:#c0392b;font-weight:bold;" if s["errors"] > 0 else "color:#888;"
        net_str   = f"{s['net_week']:+d}{trend}"

        summary_rows += f"""
        <tr>
          <td style="padding:10px 14px;border-bottom:1px solid #eee;font-weight:600;">{profile}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #eee;text-align:center;">{s['days_tracked']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #eee;text-align:center;font-weight:600;">{s['last_count']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #eee;text-align:center;{new_style}">{s['total_new']:+d}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #eee;text-align:center;{del_style}">{s['total_deleted']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #eee;text-align:center;color:{net_color};font-weight:bold;">{net_str}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #eee;text-align:center;{err_style}">{s['errors']}</td>
        </tr>"""

    daily_sections = ""
    for profile, rows in weekly_data.items():
        daily_rows = ""
        for i, r in enumerate(rows):
            bg      = "#f9f9f9" if i % 2 == 0 else "#ffffff"
            net     = r.get('net', 0)
            new     = r.get('new', 0)
            deleted = r.get('deleted', 0)
            total   = r.get('total', 0)
            
            net_str = f"{net:+d}" if net != 0 else "—"
            net_col = "#1a7a3c" if net > 0 else ("#c0392b" if net < 0 else "#888")
            new_col = "#1a7a3c" if new > 0 else "#333"
            del_col = "#c0392b" if deleted > 0 else "#333"
            daily_rows += f"""
            <tr style="background:{bg};">
              <td style="padding:8px 12px;border-bottom:1px solid #eee;">{r['date']}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;font-weight:600;">{total}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:{new_col};font-weight:{'bold' if new>0 else 'normal'};">{new:+d}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:{del_col};">{deleted}</td>
              <td style="padding:8px 12px;border-bottom:1px solid #eee;text-align:center;color:{net_col};font-weight:bold;">{net_str}</td>
            </tr>"""

        daily_sections += f"""
        <div style="margin-bottom:24px;">
          <h3 style="margin:0 0 8px 0;font-size:13px;text-transform:uppercase;
                     letter-spacing:1px;color:#2c5f8a;">{profile}</h3>
          <table width="100%" cellpadding="0" cellspacing="0"
                 style="border-collapse:collapse;font-size:13px;">
            <thead>
              <tr style="background:#2c5f8a;color:#fff;">
                <th style="padding:8px 12px;text-align:left;">Data</th>
                <th style="padding:8px 12px;text-align:center;">Ogłoszeń</th>
                <th style="padding:8px 12px;text-align:center;">Nowe</th>
                <th style="padding:8px 12px;text-align:center;">Usunięte</th>
                <th style="padding:8px 12px;text-align:center;">Netto</th>
              </tr>
            </thead>
            <tbody>{daily_rows}</tbody>
          </table>
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="pl">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:Arial,sans-serif;">
<div style="max-width:680px;margin:32px auto;background:#fff;border-radius:10px;
            overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.08);">

  <div style="background:#2c5f8a;padding:28px 32px;">
    <h1 style="margin:0;color:#fff;font-size:20px;font-weight:700;">📊 OLX Monitor</h1>
    <p style="margin:6px 0 0;color:#a8c8e8;font-size:13px;">
      Raport tygodniowy &nbsp;·&nbsp; {week_start} – {today}
    </p>
  </div>

  <div style="padding:28px 32px;">

    <h2 style="margin:0 0 16px;font-size:15px;color:#2c5f8a;text-transform:uppercase;
               letter-spacing:.5px;border-bottom:2px solid #2c5f8a;padding-bottom:8px;">
      Podsumowanie tygodnia
    </h2>
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;font-size:13px;margin-bottom:8px;">
      <thead>
        <tr style="background:#2c5f8a;color:#fff;">
          <th style="padding:10px 14px;text-align:left;">Profil</th>
          <th style="padding:10px 14px;text-align:center;">Dni</th>
          <th style="padding:10px 14px;text-align:center;">Stan</th>
          <th style="padding:10px 14px;text-align:center;">Nowe</th>
          <th style="padding:10px 14px;text-align:center;">Usun.</th>
          <th style="padding:10px 14px;text-align:center;">Netto</th>
          <th style="padding:10px 14px;text-align:center;">Błędy</th>
        </tr>
      </thead>
      <tbody>{summary_rows}</tbody>
    </table>
    <p style="margin:4px 0 24px;font-size:11px;color:#888;">
      Stan = aktualna liczba ogłoszeń &nbsp;|&nbsp; Nowe = przybyło w tygodniu &nbsp;|&nbsp;
      Usun. = usunięto &nbsp;|&nbsp; Netto = zmiana netto &nbsp;|&nbsp; Błędy = dni z błędem odczytu
    </p>

    <h2 style="margin:0 0 12px;font-size:15px;color:#2c5f8a;text-transform:uppercase;
               letter-spacing:.5px;border-bottom:2px solid #2c5f8a;padding-bottom:8px;">
      🤖 Analiza AI
    </h2>
    <div style="background:#f0f4f8;border-left:4px solid #2c5f8a;padding:16px 20px;
                border-radius:0 6px 6px 0;margin-bottom:28px;font-size:14px;
                line-height:1.7;color:#333;">
      {analysis.replace(chr(10), '<br>')}
    </div>

    <h2 style="margin:0 0 16px;font-size:15px;color:#2c5f8a;text-transform:uppercase;
               letter-spacing:.5px;border-bottom:2px solid #2c5f8a;padding-bottom:8px;">
      📅 Zestawienie dzienne
    </h2>
    {daily_sections}

  </div>

  <div style="background:#f0f4f8;padding:16px 32px;text-align:center;
              font-size:11px;color:#888;border-top:1px solid #e0e8f0;">
    Raport wygenerowany automatycznie przez OLX Monitor &nbsp;·&nbsp;
    GitHub Actions &nbsp;·&nbsp; {datetime.now().strftime("%Y-%m-%d %H:%M")}
  </div>

</div>
</body>
</html>"""


# ─── ANALIZA AI (Google Gemini) ───────────────────────────────────────────────

def generate_ai_analysis(summary: dict, weekly_data: dict) -> str:
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return "⚠ Analiza AI niedostępna – brak klucza GEMINI_API_KEY."

    data_for_ai = {}
    for profile, s in summary.items():
        data_for_ai[profile] = {
            "stan_na_koniec":   s["last_count"],
            "stan_na_poczatek": s["first_count"],
            "nowe":             s["total_new"],
            "usuniete":         s["total_deleted"],
            "zmiana_netto":     s["net_week"],
            "dni":              s["days_tracked"],
        }

    prompt = f"""Jesteś analitykiem rynku nieruchomości.
Poniżej masz tygodniowe dane z monitoringu ogłoszeń na OLX.pl (stancje i pokoje w Lublinie).

Dane z ostatnich 7 dni:
{json.dumps(data_for_ai, ensure_ascii=False, indent=2)}

Napisz zwięzłą analizę (5-8 zdań) po polsku. Uwzględnij:
- Ogólny trend na rynku pokoi w Lublinie
- Aktywność poszczególnych wynajmujących
- Czy rynek jest aktywny czy spokojny w tym tygodniu
- Krótką rekomendację dla obserwującego rynek

Pisz naturalnie, bez wypunktowań, jako spójny tekst."""

    # Próbuj modele po kolei — od najtańszego
    models = [
        "gemini-2.0-flash-lite",
        "gemini-1.5-flash-8b",
        "gemini-1.5-flash",
    ]

    for model in models:
        try:
            url  = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
            resp = requests.post(url, json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"maxOutputTokens": 500, "temperature": 0.7},
            }, timeout=30)

            if resp.status_code == 429:
                print(f"  ⚠  {model}: limit quota – próbuję kolejny model...")
                continue

            if not resp.ok:
                print(f"  ⚠  {model}: błąd {resp.status_code}")
                continue

            text = resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
            print(f"  ✓  Analiza AI wygenerowana przez {model}")
            return text

        except Exception as e:
            print(f"  ⚠  {model}: wyjątek {e}")
            continue

    return "⚠ Analiza AI chwilowo niedostępna – wszystkie modele Gemini przekroczyły limit. Spróbuj ponownie za godzinę."


# ─── WYSYŁANIE E-MAILA ───────────────────────────────────────────────────────

def send_email(subject: str, html_body: str):
    gmail_password = os.environ.get("GMAIL_APP_PASSWORD", "")
    if not gmail_password:
        print("⚠  Brak GMAIL_APP_PASSWORD – e-mail nie zostanie wysłany.")
        return False

    msg            = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = RECIPIENT_EMAIL
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    if os.path.exists(EXCEL_FILE):
        today           = datetime.now().strftime("%Y-%m-%d")
        attachment_name = f"OLX_Monitor_{today}.xlsx"
        with open(EXCEL_FILE, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=attachment_name)
        msg.attach(part)
        print(f"  📎 Załączono: {attachment_name}")
    else:
        print("  ⚠  Plik Excel nie znaleziony – wysyłam bez załącznika.")

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, gmail_password)
            server.sendmail(SENDER_EMAIL, RECIPIENT_EMAIL, msg.as_string())
        print(f"✅  E-mail wysłany → {RECIPIENT_EMAIL}")
        return True
    except Exception as e:
        print(f"❌  Błąd wysyłania e-maila: {e}")
        return False


# ─── GŁÓWNA FUNKCJA ──────────────────────────────────────────────────────────

def send_weekly_report():
    print("\n📧  Generowanie tygodniowego raportu e-mail...")

    weekly_data = get_weekly_data()
    if not weekly_data:
        print("⚠  Brak danych z ostatnich 7 dni – raport nie zostanie wysłany.")
        return

    summary  = compute_summary(weekly_data)
    analysis = generate_ai_analysis(summary, weekly_data)

    today   = datetime.now().strftime("%d.%m.%Y")
    subject = f"📊 OLX Monitor – raport tygodniowy {today}"
    html    = build_html_email(summary, weekly_data, analysis)

    send_email(subject, html)


if __name__ == "__main__":
    send_weekly_report()
